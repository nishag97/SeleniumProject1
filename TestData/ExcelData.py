import openpyxl

Dict = {}
excelbook = openpyxl.load_workbook("/Users/rws/Documents/Book1.xlsx")

current_sheet = excelbook.active

cell12 = current_sheet.cell(row=1, column=2)
print(cell12.value)

current_sheet.cell(row=2, column=2).value = "NISHI"
print(current_sheet.cell(row=2, column=2).value)
excelbook.save("/Users/rws/Documents/Book1.xlsx")

print(current_sheet.max_row)
print(current_sheet.max_column)

for i in range(1, current_sheet.max_row + 1):
    for j in range(1, current_sheet.max_column+1):
        print(current_sheet.cell(row=i, column=j).value)

#store data from excel to dictionary
for i in range(1,current_sheet.max_row+1):  # to get rows
    if current_sheet.cell(row =i,column=1).value == "T3":
        for j in range(2,current_sheet.max_column+1):
            #to get columns
            Dict[current_sheet.cell(row=1, column=j).value] = current_sheet.cell(row=i, column=j).value

print(Dict)


