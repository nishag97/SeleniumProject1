import openpyxl


class HomePageData:

    test_Homepage_data_set_1 = [{"firstname": "Dhruv", "emailId": "nishi@test.k", "gender": "Male"}]

    test_Homepage_data_set_2 = [{"firstname":"Dhruv", "emailId":"nishi@test.k", "gender":"Male"}, {"firstname":"ANJANA", "emailId":"test2e@", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("/Users/rws/Documents/Book1.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
